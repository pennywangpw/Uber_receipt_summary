import os
import json
import glob
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. 智慧判斷餐點時段 (Meal)
def get_meal_period(time_str):
    is_pm = "下午" in time_str or "PM" in time_str.upper()
    hour_match = re.search(r"(\d{1,2}):(\d{2})", time_str)
    if not hour_match:
        return "Dinner"
    hour = int(hour_match.group(1))
    if is_pm and hour != 12:
        hour += 12
    elif not is_pm and hour == 12:
        hour = 0
        
    if 5 <= hour < 11:
        return "Breakfast"
    elif 11 <= hour < 16:
        return "Lunch"
    else:
        return "Dinner"

# 2. 將中文日期轉換成 05/21/26 格式，並取得星期幾
def convert_date_and_get_weekday(date_timestamp_str):
    current_year = datetime.now().year
    year_short = str(current_year)[2:]
    match = re.search(r"(\d{1,2})月\s*(\d{1,2})日", date_timestamp_str)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        date_obj = datetime(current_year, month, day)
        return f"{month:02d}/{day:02d}/{year_short}", date_obj.strftime("%a")
    return "01/01/26", "Mon"

# 3. 核心突破：將資料夾名稱 (如 uber_receipts_2026-06-24_to_2026-06-24) 改造成分頁名稱 (06.24-06.24)
def generate_sheet_name_from_path(folder_path):
    folder_name = os.path.basename(os.path.normpath(folder_path))
    # 尋找路徑中的兩組日期特徵 (YYYY-MM-DD)
    dates = re.findall(r"(\d{4})-(\d{2})-(\d{2})", folder_name)
    if len(dates) >= 2:
        start_month, start_day = dates[0][1], dates[0][2]
        end_month, end_day = dates[1][1], dates[1][2]
        return f"{start_month}.{start_day}-{end_month}.{end_day}"
    return "UberEats 明細"

def create_new_sheet_and_fill(json_dir, excel_path):
    print(f"📂 正在處理資料夾: {json_dir}, 使用 Excel 檔案: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"❌ 找不到既有的 Excel 檔案：{excel_path}，請確認路徑。")
        return

    # 開啟既有活頁簿
    wb = openpyxl.load_workbook(excel_path)
    
    # 動態產生分頁名稱 (例如: 06.24-06.24)
    target_sheet_name = generate_sheet_name_from_path(json_dir)
    
    # 如果這個名字的分頁已經存在了，先把它刪除，避免重複追加導致資料錯亂
    if target_sheet_name in wb.sheetnames:
        del wb[target_sheet_name]
        print(f"♻️ 偵測到已存在同名分頁，已自動重置分頁: {target_sheet_name}")
        
    # 建立全新分頁
    ws = wb.create_sheet(title=target_sheet_name)
    print(f"🆕 成功創建全新分頁: {target_sheet_name}")

    # ========================================================
    # 🎨 完美還原附圖樣式：繪製雙語標頭與色彩
    # ========================================================
    # 定位顏色 (藍色與黃色)
    blue_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    yellow_fill = PatternFill(start_color="95B3D7", end_color="FFD966", fill_type="solid") # 黃色Amount
    
    font_title = Font(name="Calibri", size=11, bold=True, color="FFFFFF") # 白色粗體字
    font_amount_title = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 框線設定
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # 寫入表頭文字（完美對齊圖片換行與中英雙語）
    headers = [
        "Receipt\nNo.\n收據編號",
        "Date\n日期",
        "Day of Week",
        "Meal",
        "Name of Restaurant /\nDescription\n店家/明細",
        "Amount\n金額"
    ]
    
    ws.append(headers)
    ws.row_dimensions[1].height = 42 # 調整表頭高度以容納多行文字

    # 為表頭上色、設定字體與框線
    for col_idx in range(1, 7):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = align_center
        cell.border = thin_border
        if col_idx == 6:
            cell.fill = yellow_fill
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        else:
            cell.fill = blue_fill
            cell.font = font_title

    # ========================================================
    # 📊 讀取 JSON 並從第二列 (編號 1) 開始依序寫入資料
    # ========================================================
    json_files = glob.glob(os.path.join(json_dir, "*.json"))
    if not json_files:
        print("📁 資料夾內沒有找到任何 JSON 檔案！")
        return

    current_no = 1
    start_row = 2 # 從第二行開始寫資料

    for file_path in json_files:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            
            raw_timestamp = data.get("order_timestamp", "")
            formatted_date, day_of_week = convert_date_and_get_weekday(raw_timestamp)
            meal_period = get_meal_period(raw_timestamp)
            
            raw_amount = data.get("total_amount", "$0.00")
            cleaned_amount = float(raw_amount.replace("$", "").replace(",", "").strip())
            
            # 填入內容
            ws.cell(row=start_row, column=1, value=current_no).alignment = align_center
            ws.cell(row=start_row, column=2, value=formatted_date).alignment = align_center
            ws.cell(row=start_row, column=3, value=day_of_week).alignment = align_center
            ws.cell(row=start_row, column=4, value=meal_period).alignment = align_center
            ws.cell(row=start_row, column=5, value="Ubereats").alignment = align_center
            
            amount_cell = ws.cell(row=start_row, column=6, value=cleaned_amount)
            amount_cell.alignment = align_center
            amount_cell.number_format = '$#,##0.00' # 維持 Excel 數值原生美金格式
            
            # 為資料列加上細框線
            for c in range(1, 7):
                ws.cell(row=start_row, column=c).border = thin_border
                
            start_row += 1
            current_no += 1

    # 智慧型自動調整欄寬，防止文字太長出現 ###
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 儲存 Excel
    wb.save(excel_path)
    print(f"🎉 任務大成功！已成功在檔案中創立分頁【{target_sheet_name}】，並自動匯入 {current_no-1} 筆訂單！")

# 🎯 只有當你「單獨執行」append_to_excel.py 時這段才會跑
# 當被 main.py import 時，這段會被自動忽略，絕對不會干擾主程式！
if __name__ == "__main__":
    # 這裡放你原本寫死的測試代碼（留著平常自己單獨測試用）
    MY_EXCEL_FILE = "2026 Meal Expense.xlsx" 
    JSON_FOLDER = "./uber_receipts_2026-06-24_to_2026-06-24"
    create_new_sheet_and_fill(JSON_FOLDER, MY_EXCEL_FILE)

# # ==========================================
# # 🚀 執行設定
# # ==========================================
# MY_EXCEL_FILE = "2026 Meal Expense.xlsx" 
# JSON_FOLDER = "./uber_receipts_2026-06-24_to_2026-06-24"

# create_new_sheet_and_fill(JSON_FOLDER, MY_EXCEL_FILE)